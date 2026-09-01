"""Exportação organizada da base de prospecção para CSV e Excel."""

from __future__ import annotations

import io
from datetime import datetime
from typing import Iterable, Optional

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo
from sqlalchemy import select
from sqlalchemy.orm import Session, selectinload

from database.models import Company


SHEET_COLUMNS = {
    "Empresas": [
        "ID", "Score", "Empresa", "Razão Social", "Nome Fantasia", "CNPJ",
        "Situação Cadastral", "Tipo", "CNAE Principal", "Descrição CNAE",
        "CNAEs Secundários", "Website", "Domínio", "Telefones", "E-mails",
        "Cidade", "UF", "País", "Capital Social", "Endereço", "Status CRM",
        "Responsável", "Confiança", "Última Atualização",
    ],
    "Contatos": [
        "Empresa ID", "Empresa", "CNPJ", "Tipo", "Departamento", "Contato",
        "Verificado", "Confiança", "Fonte URL",
    ],
    "Decisores": [
        "Empresa ID", "Empresa", "CNPJ", "Nome", "Cargo", "Departamento",
        "Tipo", "Prioridade", "Score", "E-mail", "Status E-mail", "Telefone",
        "LinkedIn", "Confiança", "Fonte", "Fonte URL",
    ],
    "QSA": [
        "Empresa ID", "Empresa", "CNPJ", "Sócio / Administrador",
        "Qualificação", "País", "Representante Legal", "Fonte", "Data da Fonte",
    ],
    "Evidências": [
        "Empresa ID", "Empresa", "CNPJ", "Campo", "Valor", "Confiança",
        "Título da Fonte", "Trecho", "Fonte URL", "Coletada em",
    ],
}


class ExportService:
    def __init__(self, session: Session):
        self.session = session

    def _companies(self, min_score: int, companies: Optional[Iterable[Company]]) -> list[Company]:
        selected = list(companies) if companies is not None else None
        if selected == []:
            return []
        stmt = select(Company).options(
            selectinload(Company.contacts), selectinload(Company.decision_makers),
            selectinload(Company.partners), selectinload(Company.department_contacts),
            selectinload(Company.evidences), selectinload(Company.cnpj_matches),
        )
        if selected is not None:
            stmt = stmt.where(Company.id.in_([company.id for company in selected]))
        else:
            stmt = stmt.where(Company.score >= min_score)
        return list(self.session.scalars(stmt.order_by(Company.score.desc(), Company.updated_at.desc())).unique().all())

    def build_export_frames(
        self,
        min_score: int = 0,
        companies: Optional[Iterable[Company]] = None,
    ) -> dict[str, pd.DataFrame]:
        """Separa empresas e seus relacionamentos em abas sem duplicar o cadastro."""
        rows = {name: [] for name in SHEET_COLUMNS}

        for company in self._companies(min_score, companies):
            contacts = company.contacts
            decision_makers = company.decision_makers
            partners = company.partners
            departments = company.department_contacts
            latest_match = max(company.cnpj_matches, key=lambda item: item.id, default=None)

            phones = [item.value for item in contacts if item.contact_type in ("TELEFONE", "WHATSAPP")]
            emails = [item.value for item in contacts if "EMAIL" in item.contact_type]
            phones.extend(item.phone for item in departments if item.phone)
            phones.extend(item.whatsapp for item in departments if item.whatsapp)
            emails.extend(item.email for item in departments if item.email)

            rows["Empresas"].append({
                "ID": company.id,
                "Score": company.score,
                "Empresa": company.name,
                "Razão Social": company.legal_name or "",
                "Nome Fantasia": company.trade_name or "",
                "CNPJ": company.cnpj or "",
                "Situação Cadastral": company.status_cadastral or "",
                "Tipo": company.company_type,
                "CNAE Principal": company.cnae_code or "",
                "Descrição CNAE": company.cnae_text or "",
                "CNAEs Secundários": latest_match.cnaes_secondary if latest_match else "",
                "Website": company.website or "",
                "Domínio": company.domain,
                "Telefones": " | ".join(dict.fromkeys(phones)),
                "E-mails": " | ".join(dict.fromkeys(emails)),
                "Cidade": company.city or "",
                "UF": company.state or "",
                "País": company.country or "",
                "Capital Social": company.capital_social,
                "Endereço": latest_match.address if latest_match else "",
                "Status CRM": company.crm_status,
                "Responsável": company.assigned_to or "",
                "Confiança": company.confidence,
                "Última Atualização": company.updated_at,
            })

            for contact in contacts:
                rows["Contatos"].append({
                    "Empresa ID": company.id, "Empresa": company.name, "CNPJ": company.cnpj or "",
                    "Tipo": contact.contact_type, "Departamento": "", "Contato": contact.value,
                    "Verificado": "Sim" if contact.is_verified else "Não", "Confiança": None,
                    "Fonte URL": contact.source_url or "",
                })
            for department in departments:
                for contact_type, value in (
                    ("EMAIL_DEPARTAMENTO", department.email),
                    ("TELEFONE_DEPARTAMENTO", department.phone),
                    ("WHATSAPP_DEPARTAMENTO", department.whatsapp),
                ):
                    if value:
                        rows["Contatos"].append({
                            "Empresa ID": company.id, "Empresa": company.name, "CNPJ": company.cnpj or "",
                            "Tipo": contact_type, "Departamento": department.department, "Contato": value,
                            "Verificado": "", "Confiança": department.confidence,
                            "Fonte URL": department.source_url or "",
                        })

            for person in decision_makers:
                rows["Decisores"].append({
                    "Empresa ID": company.id, "Empresa": company.name, "CNPJ": company.cnpj or "",
                    "Nome": person.name, "Cargo": person.role, "Departamento": person.department or "",
                    "Tipo": person.decision_maker_type, "Prioridade": person.decision_priority,
                    "Score": person.decision_maker_score, "E-mail": person.email or "",
                    "Status E-mail": person.email_status, "Telefone": person.phone or "",
                    "LinkedIn": person.linkedin_url or "", "Confiança": person.confidence,
                    "Fonte": person.source_title or "", "Fonte URL": person.source_url or "",
                })

            for partner in partners:
                rows["QSA"].append({
                    "Empresa ID": company.id, "Empresa": company.name, "CNPJ": company.cnpj or "",
                    "Sócio / Administrador": partner.name, "Qualificação": partner.qualification,
                    "País": partner.country, "Representante Legal": partner.legal_representative or "",
                    "Fonte": partner.source, "Data da Fonte": partner.source_date,
                })

            for evidence in company.evidences:
                rows["Evidências"].append({
                    "Empresa ID": company.id, "Empresa": company.name, "CNPJ": company.cnpj or "",
                    "Campo": evidence.field_name, "Valor": evidence.value or "",
                    "Confiança": evidence.confidence, "Título da Fonte": evidence.source_title or "",
                    "Trecho": evidence.source_text or "", "Fonte URL": evidence.source_url or "",
                    "Coletada em": evidence.created_at,
                })

        return {
            name: pd.DataFrame(rows[name], columns=columns)
            for name, columns in SHEET_COLUMNS.items()
        }

    def build_export_dataframe(self, min_score: int = 0) -> pd.DataFrame:
        """Mantém a exportação CSV compatível, com uma linha por empresa."""
        return self.build_export_frames(min_score=min_score)["Empresas"]

    def export_to_csv(self, min_score: int = 0) -> bytes:
        return self.build_export_dataframe(min_score=min_score).to_csv(
            index=False, encoding="utf-8-sig"
        ).encode("utf-8-sig")

    def export_to_xlsx(
        self,
        min_score: int = 0,
        companies: Optional[Iterable[Company]] = None,
        scope_label: str = "Base completa",
    ) -> bytes:
        frames = self.build_export_frames(min_score=min_score, companies=companies)
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            for sheet_name, frame in frames.items():
                frame.to_excel(writer, index=False, sheet_name=sheet_name)
                self._format_sheet(writer.book[sheet_name], sheet_name, len(frame))
            writer.book.properties.title = "CXPack Radar — Empresas"
            writer.book.properties.subject = scope_label
            writer.book.properties.creator = "CXPack Radar"
            writer.book.properties.created = datetime.now()
        return output.getvalue()

    @staticmethod
    def _format_sheet(sheet, sheet_name: str, row_count: int) -> None:
        sheet.freeze_panes = "A2"
        sheet.sheet_view.showGridLines = False
        sheet.row_dimensions[1].height = 28
        for cell in sheet[1]:
            cell.fill = PatternFill("solid", fgColor="171914")
            cell.font = Font(color="F7F4EC", bold=True)
            cell.alignment = Alignment(vertical="center")

        if row_count:
            table_names = {
                "Empresas": "TabelaEmpresas", "Contatos": "TabelaContatos",
                "Decisores": "TabelaDecisores", "QSA": "TabelaQSA",
                "Evidências": "TabelaEvidencias",
            }
            table = Table(displayName=table_names[sheet_name], ref=sheet.dimensions)
            table.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False,
                showRowStripes=True, showColumnStripes=False,
            )
            sheet.add_table(table)

        headers = {cell.value: cell.column for cell in sheet[1]}
        for name in ("Confiança",):
            if name in headers:
                for cell in sheet.iter_cols(min_col=headers[name], max_col=headers[name], min_row=2):
                    for item in cell:
                        item.number_format = "0%"
        if "Capital Social" in headers:
            for row in range(2, row_count + 2):
                sheet.cell(row, headers["Capital Social"]).number_format = '#,##0.00'
        for name in ("Última Atualização", "Coletada em", "Data da Fonte"):
            if name in headers:
                for row in range(2, row_count + 2):
                    sheet.cell(row, headers[name]).number_format = "yyyy-mm-dd hh:mm"

        url_headers = {"Website", "LinkedIn", "Fonte URL"}
        for name in url_headers.intersection(headers):
            for row in range(2, row_count + 2):
                cell = sheet.cell(row, headers[name])
                if isinstance(cell.value, str) and cell.value.startswith(("http://", "https://")):
                    cell.hyperlink = cell.value
                    cell.style = "Hyperlink"

        for column in sheet.columns:
            values = [str(cell.value or "") for cell in column[: min(row_count + 1, 201)]]
            sheet.column_dimensions[column[0].column_letter].width = min(max(max(map(len, values), default=10) + 2, 12), 42)
