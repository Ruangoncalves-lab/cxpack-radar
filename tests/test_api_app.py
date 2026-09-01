import io

from fastapi.testclient import TestClient
from openpyxl import load_workbook

from api_app import app


def test_api_health():
    with TestClient(app) as client:
        response = client.get("/api/health")

    assert response.status_code == 200
    assert response.json() == {"status": "ok"}


def test_api_exports_organized_excel():
    with TestClient(app) as client:
        response = client.get("/api/companies-export.xlsx")

    assert response.status_code == 200
    assert response.headers["content-type"].startswith("application/vnd.openxmlformats")
    workbook = load_workbook(io.BytesIO(response.content), read_only=True)
    assert workbook.sheetnames == ["Empresas", "Contatos", "Decisores", "QSA", "Evidências"]
