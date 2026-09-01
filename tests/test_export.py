"""Testes da exportação organizada para Excel e CSV."""

import io
import pytest
from openpyxl import load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
from database.connection import Base
from database.models import Company, CompanyPartner, Contact, DecisionMaker, Evidence
from services.export_service import ExportService


@pytest.fixture
def db_session():
    engine = create_engine("sqlite:///:memory:")
    Base.metadata.create_all(bind=engine)
    SessionFactory = sessionmaker(bind=engine)
    session = SessionFactory()
    yield session
    session.close()


def test_export_csv_and_xlsx(db_session):
    comp = Company(
        name="Plásticos Brasil",
        domain="plasticosbrasil.com.br",
        company_type="FABRICANTE",
        score=85,
        city="São Paulo",
        state="SP"
    )
    db_session.add(comp)
    db_session.commit()
    db_session.add_all([
        Contact(company_id=comp.id, contact_type="TELEFONE", value="+55 11 3333-4444", is_verified=True, source_url="https://fonte.example/contato"),
        DecisionMaker(company_id=comp.id, name="Ana Compras", role="Gerente de Compras", email="ana@example.com", confidence=0.9, source_url="https://fonte.example/equipe"),
        CompanyPartner(company_id=comp.id, name="João Sócio", qualification="SOCIO_ADMINISTRADOR"),
        Evidence(company_id=comp.id, field_name="qualified_search_match", value="frasco PEAD", confidence=0.85, source_url="https://fonte.example/produto"),
    ])
    db_session.commit()

    exporter = ExportService(db_session)

    # Validar DataFrame
    df = exporter.build_export_dataframe()
    assert len(df) == 1
    assert df.iloc[0]["Empresa"] == "Plásticos Brasil"
    assert df.iloc[0]["Tipo"] == "FABRICANTE"

    # Validar geração de bytes CSV (UTF-8-SIG)
    csv_bytes = exporter.export_to_csv()
    assert isinstance(csv_bytes, bytes)
    assert len(csv_bytes) > 0
    assert "Plásticos Brasil".encode("utf-8") in csv_bytes

    # Validar geração de bytes XLSX (Excel)
    xlsx_bytes = exporter.export_to_xlsx()
    assert isinstance(xlsx_bytes, bytes)
    assert len(xlsx_bytes) > 0

    workbook = load_workbook(io.BytesIO(xlsx_bytes))
    assert workbook.sheetnames == ["Empresas", "Contatos", "Decisores", "QSA", "Evidências"]
    assert workbook["Empresas"].freeze_panes == "A2"
    assert workbook["Empresas"].max_row == 2
    assert workbook["Contatos"].max_row == 2
    assert workbook["Decisores"].max_row == 2
    assert workbook["QSA"].max_row == 2
    assert workbook["Evidências"].max_row == 2
    assert len(workbook["Empresas"].tables) == 1
